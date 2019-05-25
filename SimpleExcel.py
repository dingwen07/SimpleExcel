# Python 3

import openpyxl
import os

class Excel(object):
    '''Excel.

    Attributes:
        excel_file: A string indicating the path to the excel file. <class 'str'>
        activated_workseet: A integer indicating the current worksheet. <class 'int'>
        wb: An openpyxl workbook object. <class 'openpyxl.workbook.workbook.Workbook'>
        ws: An openpyxl worksheet object. <class 'openpyxl.worksheet.worksheet.Worksheet'>
    '''
    excel_file = ''
    activated_worksheet = 0

    def __init__(self, file, worksheet=0):
        '''Inits Excel with file.'''
        if not os.path.exists(file):
            raise FileNotFoundError("No such file or directory: '{}'".format(file))
        self.excel_file = file
        self.activated_worksheet = worksheet
        self.load_workbook()

    def load_workbook(self):
        self.wb = openpyxl.load_workbook(filename=self.excel_file)
        self.set_worksheet(self.activated_worksheet)

    def set_worksheet(self, worksheet):
        self.activated_worksheet = worksheet
        self.ws = self.wb.worksheets[worksheet]

    def save(self, file='default'):
        '''Save workbook to local file.
        Args:
            file: A string indicating the path to the excel file. <class 'str'>

        Returns:
            Nothing should be returned.

        Raises:
            FileNotFoundError: File not found.
            PermissionError: No write permission of this file.
        '''
        if file == 'default':
            file = self.excel_file
        elif not os.path.exists(file):
            raise FileNotFoundError("No such file or directory: '{}'".format(file))
        try:
            self.wb.save(filename=file)
        except PermissionError:
            raise

    def read(self, cell):
        '''Read the value of a spicific cell.
        Args:
            cell: A list or tuple of integers indicating the position of the cell.
                  <class 'tuple'> or <class 'list'>
                  (row, col)
                  Row or column values must be at least 1.

        Returns:
            An object indicating the value of the cell.

        Raises:
            TypeError: The argument is not valid.
            ValueError: Row or column values must be at least 1
        '''
        if not isinstance(cell, (tuple, list)) or len(cell) != 2 or type(cell[0]) != int or type(cell[1]) != int:
            raise TypeError('read() arg 2 must be a list or tuple of integers')
        return self.ws.cell(row=cell[0], column=cell[1]).value

    def read_range(self, begin, end):
        '''Read the value of all cells in a specified area
        Args:
            begin: A list or tuple of integers indicating the position of the cell.
                   <class 'tuple'> or <class 'list'>
            end: A list or tuple of integers indicating the position of the cell. 
                 <class 'tuple'> or <class 'list'>

        Returns:
            A list indicating a two-dimentianl array will be returned. <class 'list'>

        Raises:
            (From 'read' method) TypeError: The argument is not valid.
            (From 'read' method) ValueError: Row or column values must be at least 1.
        '''
        result = []
        for col in range(begin[1], end[1] + 1):
            col_result = []
            for row in range(begin[0], end[0] + 1):
                col_result.append(self.read([row, col]))
            result.append(col_result)
        return result

    def write(self, cell, value):
        '''Change the value of a spicific cell
        Args:
            cell: A list or tuple of integers indicating the position of the cell. 
                  <class 'tuple'> or <class 'list'>
            value: An object indicating the value that is intended to write. 
                   <class 'object'>

        Returns:
            An object indicating the old value of the cell. <class 'object'>

        Raises:
            ValueError: The value of the 'value' argument cannot be stored into excel file.
            (From 'read' method) ValueError: Row or column values must be at least 1.
        '''
        old_value = self.read(cell)
        try:
            self.ws.cell(row=cell[0], column=cell[1]).value = value
        except ValueError:
            try:
                self.ws.cell(row=cell[0], column=cell[1]).value = str(value)
            except ValueError:
                raise
        except:
            raise
        return old_value


    @staticmethod
    def convert(cell_name):
        '''Convert MSOffice name
        Convert MSOffice name for a cell to the format that can be reognized in this class.

        Args:
            cell_name: A string indicating the name of a cell in MSOffice. <class 'str'>
                       'A1'

        Returns:
            A list of integers indicating the position of a cell. <class 'tuple'> or <class 'list'>

        Raises:
            TypeError: The argument is not a string
        '''
        if not isinstance(cell_name, str):
            raise TypeError('The argument of convert() must be a string')
        cell_name = cell_name.upper()
        for index in range(0, len(cell_name)):
            if ord(cell_name[index]) < 65:
                colString = cell_name[0:index]
                row = int(cell_name[index:])
                break
        col = 0
        for index in range(0, len(colString)):
            col = col + (ord(colString[index]) - 64) * 26**(len(colString) - index - 1)
        return ([row, col])
